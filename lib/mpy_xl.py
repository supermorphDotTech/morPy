r"""
morPy Framework by supermorph.tech
https://github.com/supermorphDotTech

Author:     Bastian Neuwirth
Descr.:     This module delivers Microsoft Excel specific routines.

WARNING (openpyxl)
openpyxl does currently not read all possible items in an Excel file so
images and charts will be lost from existing files if they are opened
and saved with the same name.
NB you must use the English name for a function and function arguments
must be separated by commas and not other punctuation such as semi-colons.
"""

import mpy_fct
import mpy_common
import sys
import openpyxl
import openpyxl.utils.cell

from mpy_decorators import metrics, log
from openpyxl import Workbook
from openpyxl import load_workbook

@metrics
def cl_autof_rd(mpy_trace: dict, app_dict: dict, cells: list) -> dict:

    r"""
    Converts a list of cells and cell ranges to a dictionary. Overlapping cell
    ranges/cells will be autoformatted to a single reference.

    :param mpy_trace: Operation credentials and tracing information.
    :param app_dict: The morPy global dictionary containing app configurations.
    :param cells: The cell or range of cells to read from. Accepted formats:
        - Single cell: ["A1"]
        - Range of cells: ["A1:ZZ1000", "C3", ...] (letters are not case sensitive).

    :return: dict
        mpy_trace: Operation credentials and tracing.
        check: Indicates whether the function executed successfully (True/False).
        cl_dict: Dictionary where cells are keys with empty arguments:
                 {'cell1': '', 'cell2': '', ...}

    :example:
        cl_autof_rd(mpy_trace, app_dict, ["A1", "B2:C3"])
    """

    # Define operation credentials (see mpy_init.init_cred() for all dict keys)
    module = 'mpy_xl'
    operation = 'cl_autof_rd(~)'
    mpy_trace = mpy_fct.tracing(module, operation, mpy_trace)

    check = False
    cl_invalid = False
    cl_dict = {}
    rx_log = True # Logging regex functions only for debugging

    try:

        # Loop through every list item
        for cl in cells:

            # Harmonize cell letters
            cl = cl.upper()

            # Evaluate the type. If a list with 0 or more than 2 items was found, the cell
            # list is invalid. For 1 item it is a single cell and for 2 items it is a range.
            pattern = '[a-zA-Z]?[a-zA-Z]{1}[0-9]+'
            type_cl = mpy_common.regex_findall(mpy_trace, app_dict, cl, pattern, rx_log)
            type_cl_len = len(type_cl)

            # The item is a cell
            if type_cl_len == 1:

                # Add the cell to the dictionary
                cl_dict.update({type_cl[0] : ''})

                # A single cell was added to the dictionary.
                log(mpy_trace, app_dict, "debug",
                lambda: f'{app_dict["loc"]["mpy"]["cl_autof_rd_1cell"]}\n'
                        f'{app_dict["loc"]["mpy"]["cl_autof_rd_cl"]}: {type_cl[0]}')

            # The item is a range
            elif type_cl_len == 2:

                # Convert the range to a list
                # 1) Extract columns
                pattern = '[a-zA-Z]?[a-zA-Z]{1}'
                rng_col1 = mpy_common.regex_findall(mpy_trace, app_dict, type_cl[0], pattern, rx_log)
                pattern = '[a-zA-Z]?[a-zA-Z]{1}'
                rng_col2 = mpy_common.regex_findall(mpy_trace, app_dict, type_cl[1], pattern, rx_log)

                # Compare columns by string length
                if len(rng_col1) <= len(rng_col2):

                    col_from = rng_col1
                    col_to = rng_col2

                else:

                    col_from = rng_col2
                    col_to = rng_col1

                # Extract and enumerate components of columns to loop through them.
                if len(col_from[0]) == 2:

                    pattern = '[A-Z]{1}'
                    col_from = mpy_common.regex_findall(mpy_trace, app_dict, col_from, pattern, rx_log)

                    # Build the sum of columns from A to col_from for further comparison.
                    # 64 refers to the unicode value of capital A minus 1.
                    col_from_sum = abs((int(ord(col_from[0])) - 64) * 26 + (int(ord(col_from[1])) - 64))

                else:

                    col_from_sum = int(ord(col_from[0])) - 64

                # Extract and enumerate components of columns to loop through them.
                if len(col_to[0]) == 2:

                    pattern = '[A-Z]{1}'
                    col_to = mpy_common.regex_findall(mpy_trace, app_dict, col_to, pattern, rx_log)

                    # Build the sum of columns from A to col_to for further comparison.
                    # 64 refers to the unicode value of capital A minus 1.
                    col_to_sum = abs((int(ord(col_to[0])) - 64) * 26 + (int(ord(col_to[1])) - 64))

                else:

                    col_to_sum = int(ord(col_to[0])) - 64

                # Temporarily store col_from and col_to for eventual reordering
                tmp_col_from = col_from
                tmp_col_from_sum = col_from_sum
                tmp_col_to = col_to
                tmp_col_to_sum = col_to_sum

                # Compare columns by the enumerated values and exchange them if necessary
                if col_from_sum > col_to_sum:

                    col_from = tmp_col_to
                    col_from_sum = tmp_col_to_sum
                    col_to = tmp_col_from
                    col_to_sum = tmp_col_from_sum

                # 2) Extract rows
                pattern = '[0-9]+'
                rng_row1 = mpy_common.regex_findall(mpy_trace, app_dict, type_cl[0], pattern, rx_log)
                pattern = '[0-9]+'
                rng_row2 = mpy_common.regex_findall(mpy_trace, app_dict, type_cl[1], pattern, rx_log)

                # Make rows integers
                rng_row1 = int(rng_row1[0])
                rng_row2 = int(rng_row2[0])

                # Compare rows to which is higher and set start and end of rows for the range
                if rng_row1 <= rng_row2:

                    row_from = rng_row1
                    row_to = rng_row2

                else:

                    row_to = rng_row1
                    row_from = rng_row2

                # Loop through all cells and add them to the dictionary
                # 1) Loop through columns
                col_counter = col_from_sum

                while col_counter <= col_to_sum:

                    # Start from the first requested row
                    row_counter = row_from

                    # 2) Loop through rows
                    while row_counter <= row_to:

                        # Rebuild the cell
                        clmn = openpyxl.utils.cell.get_column_letter(col_counter)
                        cll = f'{clmn}{row_counter}'

                        # Add the cell to the dictionary
                        cl_dict.update({cll : ''})

                        # Iterate
                        row_counter += 1

                    # Iterate
                    col_counter += 1

                # A range of cells was added to the dictionary.
                log(mpy_trace, app_dict, "debug",
                lambda: f'{app_dict["loc"]["mpy"]["cl_autof_rd_done"]}\n'
                        f'{app_dict["loc"]["mpy"]["cl_autof_rd_rng"]}: {openpyxl.utils.cell.get_column_letter(col_from_sum)}){row_from})'
                        f' : {openpyxl.utils.cell.get_column_letter(col_to_sum)}){row_to}')

            # The item is not a valid cell
            elif type_cl_len < 1 or type_cl_len > 2:

                cl_invalid = True

                # The cell value is invalid. Autoformatting aborted.
                log(mpy_trace, app_dict, "warning",
                lambda: f'{app_dict["loc"]["mpy"]["cl_autof_rd_invalid"]}\n'
                        f'{app_dict["loc"]["mpy"]["cl_autof_rd_cls"]}: {cells}\n'
                        f'check: {check}')

        # Evaluate the validity of the dictionary
        if not cl_invalid:

            check = True

    except Exception as e:
        log(mpy_trace, app_dict, "error",
        lambda: f'{app_dict["loc"]["mpy"]["err_line"]}: {sys.exc_info()[-1].tb_lineno}\n'
                f'{app_dict["loc"]["mpy"]["err_excp"]}: {e}')

    return{
        'mpy_trace' : mpy_trace,
        'check' : check,
        'cl_dict' : cl_dict
        }

@metrics
def cl_read(mpy_trace: dict, app_dict: dict, wb_path: str, wb_sht: str, cells: list, dat: bool, vba: bool) -> dict:

    r"""
    Reads the cells of MS Excel workbooks.

    :param mpy_trace: Operation credentials and tracing information.
    :param app_dict: The morPy global dictionary containing app configurations.
    :param wb_path: Path to the MS Excel workbook.
    :param wb_sht: Name of the sheet to copy cells from.
    :param cells: The cell or range of cells to read from. Accepted formats:
        - Single cell: ["A1"]
        - Range of cells: ["A1:ZZ1000"] (letters are not case sensitive).
        This parameter is a list, and you may add multiple items.
    :param dat: If True, cells with formulae are represented by their calculated values.
                This setting requires the workbook to be reopened for changes to take effect.
    :param vba: If True, preserves any Visual Basic elements in the workbook. These elements
                remain uneditable. This setting requires the workbook to be reopened for changes.

    :return: dict
        mpy_trace: Operation credentials and tracing.
        check: Indicates whether the function executed successfully (True/False).
        cl_dict: Dictionary where cells are keys with copied arguments:
                 {'cell1': 'copied value 1', 'cell2': 'copied value 2', ...}

    :example:
        cl_read(mpy_trace, app_dict, "path/to/workbook.xlsx", "Sheet1", ["A1", "B2:C3"], True, False)
    """

    # Define operation credentials (see mpy_init.init_cred() for all dict keys)
    module = 'mpy_xl'
    operation = 'cl_read(~)'
    mpy_trace = mpy_fct.tracing(module, operation, mpy_trace)

    check = False
    cl_dict = {}
    wb_sht_active = False

    try:

        # Load the workbook and retrieve all necessary returns
        wb = wb_load(mpy_trace, app_dict, wb_path, dat, vba)
        wb_check = wb["check"]
        wb_obj = wb["wb_obj"]
        wb_path = wb["wb_path"]
        wb_sheets = wb["wb_sheets"]
        sht_active = wb["sht_active"]

        # Evaluate the availability of the workbook
        if wb_check:

            # Look up the active sheet
            if sht_active == wb_sht:

                sheet_obj = wb_obj.active
                wb_sht_active = True

                # The requested sheet is active.
                log(mpy_trace, app_dict, "debug",
                lambda: f'{app_dict["loc"]["mpy"]["cl_read_sht_active"]}\n'
                        f'{app_dict["loc"]["mpy"]["cl_read_file"]}: {wb_path}\n'
                        f'{app_dict["loc"]["mpy"]["cl_read_sht"]}: {wb_sht}')

            # Set the requested sheet active
            else:

                # The requested sheet exists as part of the workbook
                if wb_sht in wb_sheets:

                    wb_obj.active = wb_obj[wb_sht]
                    sheet_obj = wb_obj.active
                    wb_sht_active = True

                    # Activated requested sheet.
                    log(mpy_trace, app_dict, "debug",
                    lambda: f'{app_dict["loc"]["mpy"]["cl_read_sht_activated"]}\n'
                            f'{app_dict["loc"]["mpy"]["cl_read_file"]}: {wb_path}\n'
                            f'{app_dict["loc"]["mpy"]["cl_read_sht"]}: {wb_sht}')

                # The requested sheet was not found
                else:
                    # Could not find the requested workbook sheet.
                    log(mpy_trace, app_dict, "denied",
                    lambda: f'{app_dict["loc"]["mpy"]["cl_read_nfnd"]}\n'
                            f'{app_dict["loc"]["mpy"]["cl_read_file"]}: {wb_path}\n'
                            f'{app_dict["loc"]["mpy"]["cl_read_sht"]}: {wb_sht}\n'
                            f'{app_dict["loc"]["mpy"]["cl_read_av_shts"]}: {wb_sheets}')

            # Copy the requested cells
            if wb_sht_active:

                # Convert wb_sht to a standardized dictionary
                cl_dict = cl_autof_rd(mpy_trace, app_dict, cells)

                # Loop through all the cells and copy them
                for cl in cl_dict["cl_dict"]:

                    cl_dict[cl] = sheet_obj[cl].value

                check = True

                # The cells and/or ranges were copied.
                log(mpy_trace, app_dict, "debug",
                lambda: f'{app_dict["loc"]["mpy"]["cl_read_copied"]}\n'
                        f'{app_dict["loc"]["mpy"]["cl_read_file"]}: {wb_path}\n'
                        f'{app_dict["loc"]["mpy"]["cl_read_sht"]}: {wb_sht}\n'
                        f'{app_dict["loc"]["mpy"]["cl_read_cls"]}: {cells}')

    # Error detection
    except Exception as e:
        log(mpy_trace, app_dict, "error",
        lambda: f'{app_dict["loc"]["mpy"]["err_line"]}: {sys.exc_info()[-1].tb_lineno}\n'
                f'{app_dict["loc"]["mpy"]["err_excp"]}: {e}')

    return{
        'mpy_trace' : mpy_trace,
        'check' : check,
        'cl_dict' : cl_dict
        }

@metrics
def cl_write(mpy_trace: dict, app_dict: dict, wb_path: str) -> dict:

    r"""
    Writes data into cells of an Excel workbook.

    :param mpy_trace: Operation credentials and tracing information.
    :param app_dict: The morPy global dictionary containing app configurations.
    :param wb_path: Path to the MS Excel workbook.

    :return: dict
        check: Indicates whether the function executed successfully (True/False).
        mpy_trace: Operation credentials and tracing.

    :example:
        result = cl_write(mpy_trace, app_dict, "path/to/workbook.xlsx")

    TODO: Finish the function
    """

    # Define operation credentials (see mpy_init.init_cred() for all dict keys)
    module = 'mpy_xl'
    operation = 'cl_write(~)'
    mpy_trace = mpy_fct.tracing(module, operation, mpy_trace)

    check = False

    try:
        # New workbook created.
        log(mpy_trace, app_dict, "debug",
        lambda: f'{app_dict["loc"]["mpy"]["cl_write_#######"]}\n'
                f'wb_path: {wb_path}')

        check = True

    except Exception as e:
        log(mpy_trace, app_dict, "error",
        lambda: f'{app_dict["loc"]["mpy"]["err_line"]}: {sys.exc_info()[-1].tb_lineno}\n'
                f'{app_dict["loc"]["mpy"]["err_excp"]}: {e}')

    return{
        'mpy_trace' : mpy_trace,
        'check' : check
        }

@metrics
def sht_edit(mpy_trace: dict, app_dict: dict, wb_path: str, ops_dict: dict) -> dict:

    r"""
    Edits worksheets of an Excel workbook. Multiple operations can be executed
    in a single call to this function.

    :param mpy_trace: Operation credentials and tracing information.
    :param app_dict: The morPy global dictionary containing app configurations.
    :param wb_path: Path to the MS Excel workbook.
    :param ops_dict: Dictionary of operations to perform on the worksheet.
                     Only define the keys for the operations to be performed.
                     Example in order of execution:
                         {'create_sheet': ('sheet name', [INT]position),
                          'rename_sheet': ('old_name', 'new_name'),
                          'duplicate_sheet': ('source sheet', 'target sheet', [INT]position),
                          'remove_sheet': ('sheet name')}

    :return: dict
        mpy_trace: Operation credentials and tracing.
        check: Indicates whether the function executed successfully (True/False).

    :example:
        sht_edit(mpy_trace, app_dict, "path/to/workbook.xlsx", {'create_sheet': ('NewSheet', 0)})

    #TODO
        - Finish the function
    """

    # Define operation credentials (see mpy_init.init_cred() for all dict keys)
    module = 'mpy_xl'
    operation = 'sht_edit(~)'
    mpy_trace = mpy_fct.tracing(module, operation, mpy_trace)

    check = False
    wb_sht_active = False
    wb_sht = f'{wb_sht}'

    try:

        # Load the workbook and retrieve all necessary returns
        wb = wb_load(mpy_trace, app_dict, wb_path, False, True)
        wb_check = wb["check"]
        wb_obj = wb["wb_obj"]
        wb_path = wb["wb_path"]
        wb_sheets = wb["wb_sheets"]
        cnt_sheets = len(wb_sheets)
        sht_active = wb["sht_active"]

        # Evaluate the availability of the workbook
        if wb_check:

            # Evaluate if requested sheet exists as part of the workbook
            if wb_sht in wb_sheets:

                # The requested workbook sheet was found.
                log(mpy_trace, app_dict, "debug",
                lambda: f'{app_dict["loc"]["mpy"]["sht_edit_found"]}\n'
                        f'wb_sht: {wb_sht}')

            # The requested sheet was not found
            else:
                # Could not find the requested workbook sheet.
                log(mpy_trace, app_dict, "debug",
                lambda: f'{app_dict["loc"]["mpy"]["sht_edit_nfnd"]}\n'
                        f'{app_dict["loc"]["mpy"]["sht_edit_file"]}: {wb_path}\n'
                        f'{app_dict["loc"]["mpy"]["sht_edit_sht"]}: {wb_sht}\n'
                        f'{app_dict["loc"]["mpy"]["sht_edit_shts"]}: {wb_sheets}')

            # Check for and execute create_sheet
            if 'create_sheet' in ops_dict:
                    # Correction of sheet position if exceeding sheet count of the workbook.
                if ops_dict["create_sheet"](1) > cnt_sheets:

                    sht_pos = cnt_sheets

                else:

                    sht_pos = ops_dict["create_sheet"](1)

                # Evaluate if requested sheet exists as part of the workbook
                if wb_sht in wb_sheets:
                    # Operation skipped. The requested worksheet already exists.
                    log(mpy_trace, app_dict, "debug",
                    lambda: f'{app_dict["loc"]["mpy"]["sht_edit_found"]}\n'
                            f'wb_sht: {wb_sht}')

                # The requested sheet was not found
                else:
                # TODO sheet_name =
                    dummy = 0

                # Workbook sheet created.
                log(mpy_trace, app_dict, "debug",
                lambda: f'{app_dict["loc"]["mpy"]["sht_edit_create_sheet_done"]}\n'
                        f'{app_dict["loc"]["mpy"]["sht_edit_name"]}: {sheet_name}')

            # Check for and execute rename_sheet
            elif 'rename_sheet' in ops_dict:
                # Workbook sheet renamed.
                log(mpy_trace, app_dict, "debug",
                lambda: f'{app_dict["loc"]["mpy"]["sht_edit_rename_sheet_done"]}\n'
                        f'{app_dict["loc"]["mpy"]["sht_edit_old_name"]}: {old_name}\n'
                        f'{app_dict["loc"]["mpy"]["sht_edit_new_name"]}: {new_name}')

            # Check for and execute duplicate_sheet
            elif 'duplicate_sheet' in ops_dict:
                # Correction of sheet position if exceeding sheet count of the workbook.
                if ops_dict["duplicate_sheet"](2) > cnt_sheets:
                    sht_pos = cnt_sheets
                else:
                    sht_pos = ops_dict["duplicate_sheet"](2)

                    # TODO sheet_name =
                    # TODO dup_name =
                    dummy = 0

                # Workbook sheet duplicated.
                log(mpy_trace, app_dict, "debug",
                lambda: f'{app_dict["loc"]["mpy"]["sht_edit_duplicate_sheet_done"]}\n'
                        f'{app_dict["loc"]["mpy"]["sht_edit_name"]}: {sheet_name}\n'
                        f'{app_dict["loc"]["mpy"]["sht_edit_dup_name"]}: {dup_name}')

            # Check for and execute remove_sheet
            elif 'remove_sheet' in ops_dict:
                # TODO
                dummy = 0

    # Error detection
    except Exception as e:
        log(mpy_trace, app_dict, "error",
        lambda: f'{app_dict["loc"]["mpy"]["err_line"]}: {sys.exc_info()[-1].tb_lineno}\n'
                f'{app_dict["loc"]["mpy"]["err_excp"]}: {e}')

    return{
        'mpy_trace' : mpy_trace,
        'check' : check
        }

@metrics
def wb_close(mpy_trace: dict, app_dict: dict, wb_path: str) -> dict:

    r"""
    Closes an MS Excel workbook. It is recommended to use morPy functions for
    opening and closing workbooks to ensure proper handling of various cases,
    including exceptions.

    :param mpy_trace: Operation credentials and tracing information.
    :param app_dict: The morPy global dictionary containing app configurations.
    :param wb_path: Path to the MS Excel workbook.

    :return: dict
        mpy_trace: Operation credentials and tracing.
        check: Indicates whether the function executed successfully (True/False).

    :example:
        wb_close(mpy_trace, app_dict, "path/to/workbook.xlsx")
    """

    # Define operation credentials (see mpy_init.init_cred() for all dict keys)
    module = 'mpy_xl'
    operation = 'wb_close(~)'
    mpy_trace = mpy_fct.tracing(module, operation, mpy_trace)

    check = False

    try:

        # Test if a list of workbook objects was created
        if 'mpy_xl_loaded_wb_lst' in app_dict:

            # Raise an ERROR if the app_dict key exists, but is not a dictionary
            if not type(app_dict["mpy_xl_loaded_wb_lst"]) is dict:
                # The dedicated list for MS Excel workbook objects appears to be occupied by
                # a user variable (type mismatch).
                log(mpy_trace, app_dict, "error",
                lambda: f'{app_dict["loc"]["mpy"]["wb_close_lst_occupied"]}\n'
                        f'{app_dict["loc"]["mpy"]["wb_close_lst_type"]}: {type(app_dict["mpy_xl_loaded_wb_lst"])}\n'
                        f'{app_dict["loc"]["mpy"]["wb_close_lst_content"]}: {app_dict["mpy_xl_loaded_wb_lst"]}')

            # Move along if the app_dict key exists and is a dictionary
            else:

                # Check if file exists and convert the file path to a path object
                if mpy_fct.pathtool(mpy_trace, wb_path)["file_exists"]:
                    wb_path = mpy_fct.pathtool(mpy_trace, wb_path)["out_path"]

                    # Check, if the workbook is an existing object and load it, if available
                    if wb_path in app_dict["mpy_xl_loaded_wb_lst"]:

                        wb_obj = app_dict["mpy_xl_loaded_wb_lst"][wb_path]

                        # Close the workbook
                        wb_obj.close()

                        # Remove the wb_obj from the open workbooks dictionary
                        app_dict["mpy_xl_loaded_wb_lst"].pop(wb_path)

                        # The workbook object was closed.
                        log(mpy_trace, app_dict, "debug",
                        lambda: f'{app_dict["loc"]["mpy"]["wb_close_done"]}\n'
                                f'wb_path: {wb_path}\n'
                                f'wb_obj: {wb_obj}')

                    # No such workbook object exists
                    else:
                        # The workbook object could not be found.
                        log(mpy_trace, app_dict, "debug",
                        lambda: f'{app_dict["loc"]["mpy"]["wb_close_nfnd"]}\n'
                                f'wb_path: {wb_path}\n'
                                f'{app_dict["loc"]["mpy"]["wb_close_obj_loaded"]}: {app_dict["mpy_xl_loaded_wb_lst"]}')

                # No such file exists
                else:
                    # The workbook does not exist.
                    log(mpy_trace, app_dict, "debug",
                    lambda: f'{app_dict["loc"]["mpy"]["wb_close_no_wb"]}\n'
                            f'wb_path: {wb_path}')

        # No list of workbook objects was created
        else:
            # No workbook object list was created. No loaded workbooks could be found and closed.
            log(mpy_trace, app_dict, "debug",
            lambda: f'{app_dict["wb_close_no_lst"]}')

        check = True

    # Error detection
    except Exception as e:
        log(mpy_trace, app_dict, "error",
        lambda: f'{app_dict["loc"]["mpy"]["err_line"]}: {sys.exc_info()[-1].tb_lineno}\n'
                f'{app_dict["loc"]["mpy"]["err_excp"]}: {e}')

    return{
        'mpy_trace' : mpy_trace,
        'check' : check
        }

@metrics
def wb_close(mpy_trace: dict, app_dict: dict, wb_path: str) -> dict:

    r"""
    Closes an MS Excel workbook. It is recommended to use morPy functions for
    opening and closing workbooks to ensure proper handling of various cases,
    including exceptions.

    :param mpy_trace: Operation credentials and tracing information.
    :param app_dict: The morPy global dictionary containing app configurations.
    :param wb_path: Path to the MS Excel workbook.

    :return: dict
        mpy_trace: Operation credentials and tracing.
        check: Indicates whether the function executed successfully (True/False).

    :example:
        wb_close(mpy_trace, app_dict, "path/to/workbook.xlsx")
    """

    # Define operation credentials (see mpy_init.init_cred() for all dict keys)
    module = 'mpy_xl'
    operation = 'wb_close_all(~)'
    mpy_trace = mpy_fct.tracing(module, operation, mpy_trace)

    check = False

    try:

        # Test if a list of workbook objects was created
        if 'mpy_xl_loaded_wb_lst' in app_dict:
            if len(app_dict["mpy_xl_loaded_wb_lst"]) > 0:
                # Closing all open workbooks. Deleting the list.
                dyn_msg = app_dict["wb_close_all_cls_del"]
            else:
                # Deleting the list.
                dyn_msg = app_dict["wb_close_all_del"]

            # A workbook object list was found.
                log(mpy_trace, app_dict, "debug",
                lambda: f'{app_dict["loc"]["mpy"]["wb_close_all_lst_fnd"]}{dyn_msg}\n'
                        f'{app_dict["loc"]["mpy"]["wb_close_all_obj_loaded"]}: {app_dict["mpy_xl_loaded_wb_lst"]}')

            # Make a list from the list of loaded workbooks to loop through it
            wb_list = list(app_dict["mpy_xl_loaded_wb_lst"])
            i = 0

            # Loop through all workbook objects to close them
            while app_dict["mpy_xl_loaded_wb_lst"] != {}:
                # Close a workbook
                wb_close(mpy_trace, app_dict, wb_list[i])
                i += 1

            check = True

        else:
            # No workbooks list has been created. Nothing to be closed.
            log(mpy_trace, app_dict, "debug",
                lambda: f'{app_dict["wb_close_all_nothing"]}')

    # Error detection
    except Exception as e:
        log(mpy_trace, app_dict, "error",
        lambda: f'{app_dict["loc"]["mpy"]["err_line"]}: {sys.exc_info()[-1].tb_lineno}\n'
                f'{app_dict["loc"]["mpy"]["err_excp"]}: {e}')

    return{
        'mpy_trace' : mpy_trace,
        'check' : check
        }

@metrics
def wb_create(mpy_trace: dict, app_dict: dict, xl_path: str) -> dict:

    r"""
    Creates a new, empty Excel workbook at the specified path.

    :param mpy_trace: Operation credentials and tracing information.
    :param app_dict: The morPy global dictionary containing app configurations.
    :param xl_path: String or path object specifying the location of the workbook to be created.

    :return: dict
        mpy_trace: Operation credentials and tracing.
        check: Indicates whether the function executed successfully (True/False).

    :example:
        wb_create(mpy_trace, app_dict, "path/to/new_workbook.xlsx")
    """

    # Define operation credentials (see mpy_init.init_cred() for all dict keys)
    module = 'mpy_xl'
    operation = 'wb_create(~)'
    mpy_trace = mpy_fct.tracing(module, operation, mpy_trace)

    check = False

    try:

        wb = Workbook()
        wb.save(filename = f'{xl_path}')

        # New workbook created.
        log(mpy_trace, app_dict, "debug",
            lambda: f'{app_dict["loc"]["mpy"]["wb_create_done"]}\n'
                      f'xl_path: {xl_path}')

        check = True

    # Error detection
    except Exception as e:
        log(mpy_trace, app_dict, "error",
        lambda: f'{app_dict["loc"]["mpy"]["err_line"]}: {sys.exc_info()[-1].tb_lineno}\n'
                f'{app_dict["loc"]["mpy"]["err_excp"]}: {e}')

    return{
        'mpy_trace' : mpy_trace,
        'check' : check
        }

@metrics
def wb_load(mpy_trace: dict, app_dict: dict, wb_path: str, dat: bool, vba: bool) -> dict:

    r"""
    Connects to an MS Excel workbook. It is recommended to use morPy functions
    to open and close workbooks to ensure proper handling of various cases,
    including exceptions. The workbook path is linked to the object in RAM by
    creating a key in a dedicated dictionary, as shown:
        >   app_dict["mpy_xl_loaded_wb_lst"] = {wb_path: wb_obj}
    This ensures that an opened Excel workbook is not addressed twice.

    :param mpy_trace: Operation credentials and tracing information.
    :param app_dict: The morPy global dictionary containing app configurations.
    :param wb_path: Path to the MS Excel workbook.
    :param dat: If True, cells with formulae are represented by their calculated values.
                Closing and reopening the workbook is required for this behavior to change.
    :param vba: If True, preserves any Visual Basic elements in the workbook. These elements
                remain uneditable. Closing and reopening the workbook is required for this behavior to change.

    :return: dict
        mpy_trace: Operation credentials and tracing.
        check: Indicates whether the function executed successfully (True/False).
        wb_obj: The workbook object that was loaded.
        wb_path: Path-object to the MS Excel workbook.
        wb_sheets: List of the sheets in the loaded workbook.
        sht_active: Name of the active sheet after opening the workbook.

    :example:
        wb_load(mpy_trace, app_dict, "path/to/workbook.xlsx", True, False)
    """

    # Define operation credentials (see mpy_init.init_cred() for all dict keys)
    module = 'mpy_xl'
    operation = 'wb_load(~)'
    mpy_trace = mpy_fct.tracing(module, operation, mpy_trace)

    path_dct = mpy_fct.pathtool(mpy_trace, wb_path)
    wb_path = path_dct["out_path"]
    check = False
    wb_obj = 'VOID'
    wb_sheets = 'VOID'
    sht_active = 'VOID'

    try:

        # Create a list of workbook objects, if it does not yet exist
        if not 'mpy_xl_loaded_wb_lst' in app_dict:
            app_dict["mpy_xl_loaded_wb_lst"] = {}

        # Raise an ERROR if the app_dict key exists, but is not a dictionary
        elif not type(app_dict["mpy_xl_loaded_wb_lst"]) is dict:
                # The dedicated list for MS Excel workbook objects appears to be occupied by a user variable.
                log(mpy_trace, app_dict, "error",
                lambda: f'{app_dict["loc"]["mpy"]["wb_load_wblst_mismatch"]}\n'
                        f'type(app_dict[\"mpy_xl_loaded_wb_lst\"]): {type(app_dict["mpy_xl_loaded_wb_lst"])}\n'
                        f'app_dict[\"mpy_xl_loaded_wb_lst\"]: {app_dict["mpy_xl_loaded_wb_lst"]}')

        # Test if the workbook path is an existing file
        if path_dct["file_exists"]:
            if wb_path in app_dict["mpy_xl_loaded_wb_lst"]:
                # The workbook is already a loaded object. No action required.
                log(mpy_trace, app_dict, "debug",
                lambda: f'{app_dict["loc"]["mpy"]["wb_load_wb_exists"]}\n'
                        f'{app_dict["loc"]["mpy"]["wb_load_path"]}: {wb_path}\n'
                        f'{app_dict["loc"]["mpy"]["wb_load_obj"]}: {app_dict["mpy_xl_loaded_wb_lst"][wb_path]}')

                # Fetch the workbook
                wb_obj = app_dict["mpy_xl_loaded_wb_lst"][wb_path]

            else:

                # Load the workbook
                wb_obj = load_workbook(wb_path, data_only=dat, keep_vba=vba)

                # Connect the filepath to the object
                app_dict["mpy_xl_loaded_wb_lst"][wb_path] = wb_obj

                # MS Excel workbook loaded.
                log(mpy_trace, app_dict, "debug",
                lambda: f'{app_dict["loc"]["mpy"]["wb_load_wb_loaded"]}\n'
                        f'{app_dict["loc"]["mpy"]["wb_load_path"]}: {wb_path}')

            # Retrieve the name of the active sheet
            sht_active = app_dict["mpy_xl_loaded_wb_lst"][wb_path].active.title

            # Retrieve all sheetnames of the workbook
            wb_sheets = app_dict["mpy_xl_loaded_wb_lst"][wb_path].sheetnames

            check = True

        else:
            # The file does not exist.
            log(mpy_trace, app_dict, "denied",
            lambda: f'{app_dict["loc"]["mpy"]["wb_load_no_file"]}\n'
                    f'{app_dict["loc"]["mpy"]["wb_load_path"]}: {wb_path}\n'
                    f'file_exists: {path_dct["file_exists"]}\n'
                    f'file_name: {path_dct["file_name"]}\n'
                    f'dir_exists: {path_dct["dir_exists"]}\n'
                    f'dir_name: {path_dct["dir_name"]}')

    except Exception as e:
        log(mpy_trace, app_dict, "error",
        lambda: f'{app_dict["loc"]["mpy"]["err_line"]}: {sys.exc_info()[-1].tb_lineno}\n'
                f'{app_dict["loc"]["mpy"]["err_excp"]}: {e}')

    return{
        'mpy_trace' : mpy_trace,
        'check' : check,
        'wb_obj' : wb_obj,
        'wb_path' : wb_path,
        'wb_sheets' : wb_sheets,
        'sht_active' : sht_active
        }

@metrics
def wb_tbl_attributes(mpy_trace: dict, app_dict: dict, wb_path: str, tbl: str) -> dict:

    r"""
    Retrieves all attributes of an MS Excel table. This function uses the
    openpyxl module to gather detailed table attributes. For basic parameters
    of an MS Excel table, consider using `wb_tbl_inquiry` instead.

    :param mpy_trace: Operation credentials and tracing information.
    :param app_dict: The morPy global dictionary containing app configurations.
    :param wb_path: Path to the MS Excel workbook.
    :param tbl: Name of the table to be analyzed.

    :return: dict
        mpy_trace: Operation credentials and tracing.
        check: Indicates whether the function executed successfully (True/False).
        wb_obj: The workbook object that was loaded.
        tbl_attr: List of the table's attributes.

    :example:
        wb_tbl_attributes(mpy_trace, app_dict, "path/to/workbook.xlsx", "Table1")
    """

    # Define operation credentials (see mpy_init.init_cred() for all dict keys)
    module = 'mpy_xl'
    operation = 'wb_tbl_attributes(~)'
    mpy_trace = mpy_fct.tracing(module, operation, mpy_trace)

    check = False
    wb_path = mpy_fct.pathtool(mpy_trace, wb_path)["out_path"]
    wb_obj = 'VOID'
    tbl_attr = 'VOID'

    try:
        # Load the workbook and inquire for tables
        wb = wb_tbl_inquiry(mpy_trace, app_dict, wb_path)

        # Retrieve the workbook object
        wb_obj = app_dict["mpy_xl_loaded_wb_lst"][wb_path]

        # Inquire the according worksheet of the table
        rtrn = wb["tbl_sht"]
        sht = rtrn[tbl]

        # Get all values of an MS Excel table
        datb_temp = wb_obj[sht].tables.values()
        datb_temp = opyxl_tbl_datb_dict(mpy_trace, app_dict, datb_temp, tbl)
        tbl_attr = datb_temp["tbl_attr"]

        # Retrieved all values of an MS Excel table.
        log(mpy_trace, app_dict, "debug",
        lambda: f'{app_dict["loc"]["mpy"]["wb_tbl_attributes_retr"]}\n'
                f'{app_dict["loc"]["mpy"]["wb_tbl_attributes_path"]}: {wb_path}\n'
                f'{app_dict["loc"]["mpy"]["wb_tbl_attributes_sheet"]}: {sht}\n'
                f'{app_dict["loc"]["mpy"]["wb_tbl_attributes_tbl"]}: {tbl}')

        check = True

    except Exception as e:
        log(mpy_trace, app_dict, "error",
        lambda: f'{app_dict["loc"]["mpy"]["err_line"]}: {sys.exc_info()[-1].tb_lineno}\n'
                f'{app_dict["loc"]["mpy"]["err_excp"]}: {e}')

    return{
        'mpy_trace' : mpy_trace,
        'check' : check,
        'wb_obj' : wb_obj,
        'tbl_attr' : tbl_attr
        }

@metrics
def wb_tbl_inquiry(mpy_trace: dict, app_dict: dict, wb_path: str) -> dict:

    r"""
    Inquires all tables in a worksheet and returns a dictionary with details
    about the tables and their attributes.

    :param mpy_trace: Operation credentials and tracing information.
    :param app_dict: The morPy global dictionary containing app configurations.
    :param wb_path: Path to the MS Excel workbook.

    :return: dict
        mpy_trace: Operation credentials and tracing.
        check: Indicates whether the function executed successfully (True/False).
        wb_obj: The workbook object that was loaded.
        wb_sheets: List of the sheets in the loaded workbook.
        tbl_lst: List of tables in the entire workbook.
        tbl_rng: List of tuples of tables and their ranges.
        tbl_sht: List of tuples of tables and the sheets they are on.

    :example:
        wb_tbl_inquiry(mpy_trace, app_dict, "path/to/workbook.xlsx")
    """

    # Define operation credentials (see mpy_init.init_cred() for all dict keys)
    module = 'mpy_xl'
    operation = 'wb_tbl_inquiry(~)'
    mpy_trace = mpy_fct.tracing(module, operation, mpy_trace)

    check = False
    wb_path = mpy_fct.pathtool(mpy_trace, wb_path)["out_path"]
    wb_obj = 'VOID'
    void_dict = {'VOID' : 'VOID'}
    tbl_rng = void_dict
    tbl_sht = void_dict
    void_list = ["VOID"]
    wb_sheets = ["VOID"]
    tbl_lst = void_list
    dat = False
    vba = False

    try:

        # Load the workbook
        wb = wb_load(mpy_trace, app_dict, wb_path, dat, vba)

        # Retrieve the workbook object
        wb_obj = app_dict["mpy_xl_loaded_wb_lst"][wb_path]

        # Retrieve all sheet names of the workbook
        wb_sheets = wb["wb_sheets"]

        # Retrieve the tables dictionary and create a dictionary with the table
        # being the key and the sheet of the workbook being the object
        for sht in wb_sheets:

            # Get all tables of the sheet "sht" including the range (List)
            tbl_tmp = wb_obj[sht].tables.items()

            # Build the dictionaries
            for index, tuple in enumerate(tbl_tmp):

                # Extract table and range from the List of tuples
                tbl = tuple[0]
                rng = tuple[1]

                # Create a list of tables
                if tbl_lst == void_list:
                    tbl_lst = [tbl]
                else:
                    tbl_lst.append(tbl)

                # Create a dictionary of tables and ranges
                if tbl_rng == void_dict:
                    tbl_rng = {tbl : rng}
                else:
                    tbl_rng.update({tbl : rng})

                # Create a dictionary of tables and sheets
                if tbl_sht == void_dict:
                    tbl_sht = {tbl : sht}
                else:
                    tbl_sht.update({tbl : sht})

        # Retrieved all tables of a workbook.
        log(mpy_trace, app_dict, "debug",
        lambda: f'{app_dict["loc"]["mpy"]["wb_tbl_inquiry_retr"]}\n'
                f'{app_dict["loc"]["mpy"]["wb_tbl_inquiry_Path"]}: {wb_path}\n\n'
                f'{app_dict["loc"]["mpy"]["wb_tbl_inquiry_tbl_wksh"]}:\n{tbl_sht}\n\n'
                f'{app_dict["loc"]["mpy"]["wb_tbl_inquiry_tbl_rng"]}:\n{tbl_rng}')

        check = True

    except Exception as e:
        log(mpy_trace, app_dict, "error",
        lambda: f'{app_dict["loc"]["mpy"]["err_line"]}: {sys.exc_info()[-1].tb_lineno}\n'
                f'{app_dict["loc"]["mpy"]["err_excp"]}: {e}')

    return{
        'mpy_trace' : mpy_trace,
        'check' : check,
        'wb_obj' : wb_obj,
        'wb_sheets' : wb_sheets,
        'tbl_lst' : tbl_lst,
        'tbl_rng' : tbl_rng,
        'tbl_sht' : tbl_sht
        }

@metrics
def opyxl_tbl_datb_dict(mpy_trace: dict, app_dict: dict, datb: object, tbl: str) -> dict:

    r"""
    Converts the interface of an openpyxl databook into a dictionary containing
    all attributes of the specified table. This function is a helper and is
    typically called by `wb_tbl_inquiry` for ease of coding.

    :param mpy_trace: Operation credentials and tracing information.
    :param app_dict: The morPy global dictionary containing app configurations.
    :param datb: The databook object as generated by openpyxl.
    :param tbl: Name of the table to be analyzed.

    :return: dict
        mpy_trace: Operation credentials and tracing.
        check: Indicates whether the function executed successfully (True/False).
        tbl_attr: List containing all attributes of the openpyxl databook.

    :example:
        opyxl_tbl_datb_dict(mpy_trace, app_dict, databook_obj, "Table1")
    """

    # Define operation credentials (see mpy_init.init_cred() for all dict keys)
    module = 'mpy_xl'
    operation = 'tbl_opyxl_datb_dict(~)'
    mpy_trace = mpy_fct.tracing(module, operation, mpy_trace)

    check = False
    datb = f'{datb}'

    try:
        # Search for regular expressions in the databook to extract only the
        # relevant Part.

        # 1. Purge all whitespace characters to make regex easier and more precise
        datb = mpy_common.regex_replace(mpy_trace, app_dict, datb, '\s', '')

        # 2. Split the databook into a list of distinct table attributes
        delimiter = '<openpyxl.worksheet.table.Tableobject>'
        datb_list = mpy_common.regex_split(mpy_trace, app_dict, datb, delimiter)

        # 3. Iterate through the list in search for the table and delete
        # elements not associated with it
        pattern = f'\'{tbl}\''

        for item in datb_list:

            result = mpy_common.regex_find1st(mpy_trace, app_dict, item, pattern)

            if result:
                break

        # 4. Replace the comma at the end of the string, if there is any
        item = mpy_common.regex_replace(mpy_trace, app_dict, item, ',$', '')

        # 5. Add the first delimiter and reinsert some spaces for compatibility
        # with Openpyxl
        item = f'<openpyxl.worksheet.table.Tableobject>{item}'
        item = mpy_common.regex_replace(mpy_trace, app_dict, item, 'object>', ' object>')

        # 6. Split the string into different sections
        tbl_attr = mpy_common.regex_split(mpy_trace, app_dict, item, ',')

        # Converted an openpyxl databook into a list specific to the attributes of the MS Excel table.
        log(mpy_trace, app_dict, "debug",
        lambda: f'{app_dict["loc"]["mpy"]["opyxl_tbl_datb_dict_conv"]}\n'
                f'{app_dict["loc"]["mpy"]["opyxl_tbl_datb_dict_tbl"]}: {tbl}\n'
                f'{app_dict["loc"]["mpy"]["opyxl_tbl_datb_dict_attr"]}:\n{tbl_attr}')

        check = True

    except Exception as e:
        log(mpy_trace, app_dict, "error",
        lambda: f'{app_dict["loc"]["mpy"]["err_line"]}: {sys.exc_info()[-1].tb_lineno}\n'
                f'{app_dict["loc"]["mpy"]["err_excp"]}: {e}')

        return{
            'mpy_trace' : mpy_trace,
            'check' : check,
            'tbl_attr' : tbl_attr
            }