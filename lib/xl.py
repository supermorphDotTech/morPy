r"""
morPy Framework by supermorph.tech
https://github.com/supermorphDotTech

Author:     Bastian Neuwirth
Descr.:     This module delivers Microsoft Excel specific routines.

NOTES on OpenPyXL:
OpenPyXL does currently not read all possible items in an Excel file, so
images and charts will be lost from existing files if they are opened
and saved with the same name.
Also, you must use the English name for a function and function arguments
must be separated by commas and not other punctuation such as semicolons.
"""

import lib.fct as morpy_fct
import lib.common as common
from morPy import log
from lib.decorators import morpy_wrap

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection
import openpyxl.utils.cell


class XlWorkbook:
    r"""
    Instantiates an Excel workbook to edit, read from and write to it.
    It uses OpenPyXL as the API to the workbooks.
    """


    __slots__ = (
        "wb_obj",
        "wb_path",
        "wb_sheets",
        "active_sheet",
        "active_sheet_title",
        "tables_ranges",
        "tables_sheets",
        "file_ext",
        "files_vbs",
    )


    @morpy_wrap
    def __init__(self, trace: dict, app_dict: dict, workbook: str, create: bool=False,
              data_only: bool=False, keep_vba: bool=True) -> None:
        r"""
        Class constructor.

        :param trace: operation credentials and tracing information
        :param app_dict: morPy global dictionary containing app configurations
        :param workbook: Path of the workbook
        :param create: If True and file does not yet exist, will create the workbook. If file exists, will
                    open the existing file.
        :param data_only: If True, cells with formulae are represented by their calculated values.
                    Closing and reopening the workbook is required for this to change.
        :param keep_vba: If True, preserves any Visual Basic elements in the workbook. Only has an effect
                    on workbooks supporting vba (i.e. xlsm-files). These elements remain immutable. Closing
                    and reopening the workbook is required to change behaviour.

        :example:
            import morPy
            wb_path = "C:\my_file.xlsx"
            
            # Create an Excel workbook
            wb = morPy.XlWorkbook(trace, app_dict, wb_path, create=True)
        """

        self.wb_path = morpy_fct.pathtool(workbook)["out_path"]

        # Opening MS Excel workbook.
        log(trace, app_dict, "info",
            lambda: f'{app_dict["loc"]["morpy"]["XlWorkbook_construct"]}\n'
                    f'{app_dict["loc"]["morpy"]["XlWorkbook_wb"]}: {self.wb_path}')

        self.wb_sheets = []
        self.active_sheet = None
        self.active_sheet_title = ""
        self.tables_ranges = {}
        self.tables_sheets = {}

        # Get the file extension in lowercase
        self.file_ext = morpy_fct.pathtool(self.wb_path)["file_ext"].lower()

        # Set vba supporting file extensions
        self.files_vbs = {".xlsm", ".xltm"}

        path_eval = morpy_fct.pathtool(self.wb_path)
        file_exists = path_eval["file_exists"]

        if not file_exists:
            if create:
                self._create_workbook(trace, app_dict)
            else:
                # File does not exist and was not created.
                raise LookupError(
                    f'{app_dict["loc"]["morpy"]["XlWorkbook_not_create"]}\n'
                    f'{app_dict["loc"]["morpy"]["XlWorkbook_create"]}: {create}'
                )

        # Re-evaluate file creation
        path_eval = morpy_fct.pathtool(self.wb_path)
        is_file = path_eval["is_file"]
        if not is_file:
            # The path to the workbook is invalid.
            raise LookupError(
                f'{app_dict["loc"]["morpy"]["XlWorkbook_path_invalid"]}\n'
                f'{app_dict["loc"]["morpy"]["XlWorkbook_path"]}: {self.wb_path}'
            )

        # Evaluate, if file type supports vba
        if keep_vba and self.file_ext not in self.files_vbs:
            keep_vba = False

        # Open the workbook
        self.wb_obj = load_workbook(self.wb_path, data_only=data_only, keep_vba=keep_vba)

        # Update workbook metadata
        self._update_meta(trace, app_dict)

        # MS Excel workbook instantiated.
        log(trace, app_dict, "debug",
            lambda: f'{app_dict["loc"]["morpy"]["XlWorkbook_inst"]}'
                    f'{app_dict["loc"]["morpy"]["XlWorkbook_wb"]}: {self.wb_path}')


    @morpy_wrap
    def _create_workbook(self, trace: dict, app_dict: dict) -> None:
        r"""
        Creates a new, empty Excel workbook at the specified path.

        :param trace: Operation credentials and tracing information.
        :param app_dict: The morPy global dictionary containing app configurations.

        :example:
            self._create_workbook(trace, app_dict)
        """

        wb = Workbook()
        wb.save(filename=f'{self.wb_path}')

        # MS Excel workbook created.
        log(trace, app_dict, "debug",
            lambda: f'{app_dict["loc"]["morpy"]["create_workbook_done"]}\n'
                    f'xl_path: {self.wb_path}')


    # Suppress linting for mandatory arguments.
    # noinspection PyUnusedLocal
    @morpy_wrap
    def _update_meta(self, trace: dict, app_dict: dict, minimal: bool=False) -> None:
        r"""
        Update which metadata of the workbook. This could be which sheet is active
        and which sheets there are.

        :param trace: Operation credentials and tracing information.
        :param app_dict: The morPy global dictionary containing app configurations.
        :param minimal: If True, updates only the attributes that could change
            without actually writing to the cell (i.e. active_sheet).
        """

        table_range = None
        table_sheet = None
        table_list = None

        # Retrieve the name of the active sheet
        self.active_sheet = self.wb_obj.active
        self.active_sheet_title = self.wb_obj.active.title

        if not minimal:
            # Retrieve all sheet names of the workbook
            self.wb_sheets = self.wb_obj.sheetnames

            # Retrieve the tables dictionary and create a dictionary with the table
            # being the key and the sheet of the workbook being the object
            for sheet in self.wb_sheets:
                # Get all tables of the sheet including the range (List)
                table_data = self.wb_obj[sheet].tables.items()

                # Build the dictionaries
                for tpl in table_data:
                    table = tpl[0]
                    table_range = tpl[1]

                    # Create a list of tables
                    if not table_list:
                        table_list = [table]
                    else:
                        table_list.append(table)

                    # Create a dictionary of tables and ranges
                    table_range = {table: table_range} if not isinstance(table_range, dict) else table_range.update({table: table_range})

                    # Create a dictionary of tables and sheets
                    table_sheet = {table: sheet} if not isinstance(table_sheet, dict) else table_sheet.update({table: sheet})

            # Store table ranges table sheets
            self.tables_ranges = table_range if table_range else {}
            self.tables_sheets = table_sheet if table_sheet else {}


    @morpy_wrap
    def _cell_ref_autoformat(self, trace: dict, app_dict: dict, cell_range: list) -> dict:
        r"""
        Converts a list of cells and cell ranges to a dictionary. Overlapping cell
        ranges/cells will be auto-formatted to a single reference.
    
        :param trace: Operation credentials and tracing information.
        :param app_dict: The morPy global dictionary containing app configurations.
        :param cell_range: The cell or range of cells to read from. Accepted formats:
            - Single cell: ["A1"]
            - Range of cells: ["A1:ZZ1000", "C3", ...] (not case-sensitive).
    
        :return: dict
            cl_dict: Dictionary where cells are keys with empty arguments:
                     {'cell1': '', 'cell2': '', ...}
    
        :example:
            cl_range = ["A1", "B2:C3", "B1:ZZ5"]
            cl_dict = self._cell_ref_autoformat(trace, app_dict, cell_range = cl_range)["cl_dict"]
        """
    
        cl_valid = False
        cl_dict = {}
    
        # Loop through every list item
        for cl in cell_range:

            # Harmonize cell letters
            cl = cl.upper()

            # Evaluate the type. If a list with 0 or more than 2 items was found, the cell
            # list is invalid. For 1 item it is a single cell and for 2 items it is a range.
            pattern = '[a-zA-Z]?[a-zA-Z]{1}[0-9]+'
            type_cl = common.regex_findall(trace, app_dict, cl, pattern)["result"]
            type_cl_len = len(type_cl)

            # The item is a cell
            if type_cl_len == 1:
                # Add the cell to the dictionary
                cl_dict.update({type_cl[0] : ''})

                # A single cell was added to the dictionary.
                log(trace, app_dict, "debug",
                    lambda: f'{app_dict["loc"]["morpy"]["cell_ref_autoformat_1cell"]}\n'
                            f'{app_dict["loc"]["morpy"]["cell_ref_autoformat_cl"]}: {type_cl[0]}')

                cl_valid = True

            # The item is a range
            elif type_cl_len == 2:
                # Convert the range to a list
                # 1) Extract columns
                pattern = '[a-zA-Z]?[a-zA-Z]{1}'
                range_col1 = common.regex_findall(trace, app_dict, type_cl[0], pattern)["result"]
                pattern = '[a-zA-Z]?[a-zA-Z]{1}'
                range_col2 = common.regex_findall(trace, app_dict, type_cl[1], pattern)["result"]

                # Compare columns by string length
                if len(range_col1) <= len(range_col2):
                    col_from = range_col1
                    col_to = range_col2
                else:
                    col_from = range_col2
                    col_to = range_col1

                # Extract and enumerate components of columns to loop through them.
                if len(col_from[0]) == 2:
                    pattern = '[A-Z]{1}'
                    col_from = common.regex_findall(trace, app_dict, col_from, pattern)["result"]

                    # Build the sum of columns from A to col_from for further comparison.
                    # 64 refers to the Unicode value of capital A minus 1.
                    col_from_sum = abs((int(ord(col_from[0])) - 64) * 26 + (int(ord(col_from[1])) - 64))
                else:
                    col_from_sum = int(ord(col_from[0])) - 64

                # Extract and enumerate components of columns to loop through them.
                if len(col_to[0]) == 2:
                    pattern = '[A-Z]{1}'
                    col_to = common.regex_findall(trace, app_dict, col_to, pattern)["result"]

                    # Build the sum of columns from A to col_to for further comparison.
                    # 64 refers to the Unicode value of capital A minus 1.
                    col_to_sum = abs((int(ord(col_to[0])) - 64) * 26 + (int(ord(col_to[1])) - 64))
                else:
                    col_to_sum = int(ord(col_to[0])) - 64

                # Temporarily store col_from and col_to for eventual reordering
                tmp_col_from_sum = col_from_sum
                tmp_col_to_sum = col_to_sum

                # Compare columns by the enumerated values and exchange them if necessary
                if col_from_sum > col_to_sum:
                    col_from_sum = tmp_col_to_sum
                    col_to_sum = tmp_col_from_sum

                # 2) Extract rows
                pattern = '[0-9]+'
                range_row1 = common.regex_findall(trace, app_dict, type_cl[0], pattern)["result"]
                pattern = '[0-9]+'
                range_row2 = common.regex_findall(trace, app_dict, type_cl[1], pattern)["result"]

                # Make rows integers
                range_row1 = int(range_row1[0])
                range_row2 = int(range_row2[0])

                # Compare rows to which is higher and set start and end of rows for the range
                if range_row1 <= range_row2:
                    row_from = range_row1
                    row_to = range_row2
                else:
                    row_to = range_row1
                    row_from = range_row2

                # Loop through all cells and add them to the dictionary
                # 1) Loop through columns
                col_counter = col_from_sum

                while col_counter <= col_to_sum:
                    # Start from the first requested row
                    row_counter = row_from

                    # 2) Loop through rows
                    while row_counter <= row_to:
                        # Rebuild the cell
                        column = openpyxl.utils.cell.get_column_letter(col_counter)
                        cll = f'{column}{row_counter}'

                        # Add the cell to the dictionary
                        cl_dict.update({cll : ''})

                        # Iterate
                        row_counter += 1

                    # Iterate
                    col_counter += 1

                # A range of cells was added to the dictionary.
                log(trace, app_dict, "debug",
                    lambda: f'{app_dict["loc"]["morpy"]["cell_ref_autoformat_done"]}\n'
                            f'{app_dict["loc"]["morpy"]["cell_ref_autoformat_rng"]}: '
                            f'({openpyxl.utils.cell.get_column_letter(col_from_sum)}:{row_from}) - '
                            f'({openpyxl.utils.cell.get_column_letter(col_to_sum)}:{row_to})')

                cl_valid = True

            # The item is not a valid cell
            else:
                cl_valid = False

                # The cell value is invalid. Autoformatting aborted.
                log(trace, app_dict, "warning",
                    lambda: f'{app_dict["loc"]["morpy"]["cell_ref_autoformat_invalid"]}\n'
                            f'{app_dict["loc"]["morpy"]["cell_ref_autoformat_cls"]}: {cell_range}\n'
                            f'check: {check}')

        # Evaluate the validity of the dictionary
        if cl_valid:
            check: bool = True
    
        return{
            'cl_dict' : cl_dict
            }


    @morpy_wrap
    def save_workbook(self, trace: dict, app_dict: dict, close_workbook: bool=False) -> None:
        r"""
        Saves the changes to the MS Excel workbook.

        :param trace: Operation credentials and tracing information.
        :param app_dict: The morPy global dictionary containing app configurations.
        :param close_workbook: If True, closes the workbook.

        :example:
            import morPy
            wb_path = "C:\my_file.xlsx"
            wb = morPy.XlWorkbook(trace, app_dict, wb_path)
            
            # Save and close the workbook. Write "None" to the reference "wb".
            wb.save_workbook(trace, app_dict, close=True)
        """

        # Save the workbook
        self.wb_obj.save(filename=self.wb_path)

        if close_workbook:
            # Close the workbook
            self.close_workbook(trace, app_dict)


    @morpy_wrap
    def close_workbook(self, trace: dict, app_dict: dict) -> None:
        r"""
        Closes the MS Excel workbook.

        :param trace: Operation credentials and tracing information.
        :param app_dict: The morPy global dictionary containing app configurations.

        :example:
            import morPy
            wb_path = "C:\my_file.xlsx"
            wb = morPy.XlWorkbook(trace, app_dict, wb_path, create=True)
            
            # Close the workbook. Write "None" to the reference "wb".
            wb.close_workbook(trace, app_dict)
        """


        # Close the workbook
        wb_path = self.wb_path
        self.wb_obj.close()

        # The workbook object was closed.
        log(trace, app_dict, "debug",
            lambda: f'{app_dict["loc"]["morpy"]["close_workbook_done"]}\n'
                    f'{app_dict["loc"]["morpy"]["close_workbook_path"]}: {wb_path}')


    @morpy_wrap
    def activate_worksheet(self, trace: dict, app_dict: dict, worksheet: str) -> None:
        r"""
        Activates a specified worksheet in the workbook. If the sheet is not found,
        an error is logged.

        :param trace: Operation credentials and tracing information.
        :param app_dict: The morPy global dictionary containing app configurations.
        :param worksheet: The name of the worksheet to activate.

        :example:
            import morPy
            wb_path = "C:\my_file.xlsx"
            w_sht = "Sheet1"
            wb = morPy.XlWorkbook(trace, app_dict, wb_path)
            wb.activate_worksheet(trace, app_dict, w_sht)
        """

        # Check if the requested sheet exists in the workbook
        if worksheet in self.wb_sheets:
            # Check for the active sheet
            self._update_meta(trace, app_dict, minimal=True)

            if not worksheet == self.active_sheet_title:
                # Set the requested sheet as active
                self.wb_obj.active = self.wb_obj[worksheet]
                # Store active sheet in meta data
                self._update_meta(trace, app_dict, minimal=True)

                # The worksheet was successfully activated.
                log(trace, app_dict, "debug",
                    lambda: f'{app_dict["loc"]["morpy"]["activate_worksheet_done"]}\n'
                            f'{app_dict["loc"]["morpy"]["activate_worksheet_sht"]}: {worksheet}')

        else:
            # The requested sheet was not found.
            raise ValueError(
                f'{app_dict["loc"]["morpy"]["activate_worksheet_not_found"]}\n'
                f'{app_dict["loc"]["morpy"]["activate_worksheet_file"]}: {self.wb_path}\n'
                f'{app_dict["loc"]["morpy"]["activate_worksheet_req_sht"]}: {worksheet}'
            )


    @morpy_wrap
    def read_cells(self, trace: dict, app_dict: dict, cell_range: list=None,
                   cell_styles: bool=False, worksheet: str=None, gui=None) -> dict:
        r"""
        Reads the cells of MS Excel workbooks. Overlapping ranges will get auto-formatted
        to ensure every cell is addressed only once.

        :param trace: Operation credentials and tracing information.
        :param app_dict: The morPy global dictionary containing app configurations.
        :param cell_range: The cell or range of cells to read from. If cell range is None, will read all cells
            inside the matrix of max_row and max_column (also empty ones). formats:
            - not case-sensitive
            - Single cell: ["A1"]
            - Range of cells: ["A1:ZZ1000"]
            - Ranges of cells: ["A1:ZZ1000", "c2:fl342"]
        :param cell_styles: If True, cell styles will be retrieved. If False, get value only.
        :param worksheet: Name of the worksheet, where the cell is located. If None, the
            active sheet is addressed.
        :param gui: User Interface reference. Automatically referenced by morPy.ProgressTrackerTk()

        :return: dict
            cl_dict: Dictionary of cell content dictionaries containing values and styles of cells. Following is
                    an example of a single complete cell write:
                cl_dict = {
                    "A1" : {
                        "value" : "Data in Cell",
                        "comment" : {
                            "text" : "This is a comment",
                            "author" : "Mr. Author Man",
                        },
                        "format" : "General",           # Options: General|Number|Currency|Accounting|Date|Time|Percentage
                                                        # > Fraction|Scientific|Text|custom strings
                        "font" : {
                            "name" : "Calibri",
                            "bold" : True,
                            "italic" : True,
                            "vertical align" : None,    # Options: None|superscript|subscript
                            "underline" : None,         # Options: None|single|double|singleAccounting|doubleAccounting
                            "strike" : False,
                            "size" : 14,
                            "color" : "D1760C",
                        },
                        "background" : {
                            "fill type" : None,         # Options: None|solid|darkGrid|darkTrellis|lightDown|lightGray
                                                        # > lightGrid|lightHorizontal|lightTrellis|lightUp
                                                        # > lightVertical|mediumGray
                            "start color" : "307591",
                            "end color" : "67CCCC",
                        },
                        "border" : {
                            "edge" : "outline",         # Options: left|right|top|bottom|diagonal|outline
                                                        # > vertical|horizontal
                            "style" : None,             # Options: None|dashDot|dashDotDot|dashed|dotted|double|hair|medium
                                                        # > medium|mediumDashDot|mediumDashDotDot|mediumDashed|slantDashDot
                                                        # > thick|thin
                            "color" : "2A7F7F",
                            "diagonal direction" : 0,   # Options: 0 (no diagonal) | 1 (downwards) | 2 (upwards)
                        },
                        "alignment" : {
                            "horizontal" : "general",   # Options: general|left|center|right|fill|justify|centerContinuous
                                                        # > distributed
                            "vertical" : "center",      # Options: top|center|bottom|justify|distributed
                            "text rotation" : 0,
                            "wrap text" : False,
                            "shrink to fit" : False,
                            "indent" : 0,
                        },
                        "protection" : {
                            "locked" : False,
                            "hidden" : False,
                        },
                    },
                    "A2" : { ... }
                }

        :example:
            import morPy
            wb_path = "C:\my_file.xlsx"
            wb = morPy.XlWorkbook(trace, app_dict, wb_path)
            range = ["A1", "B2:C3", "B1:ZZ5"]
            cell_dict = wb.read_cells(trace, app_dict, worksheet="Sheet1", cell_range=range)["cl_dict"]
        """

        # Update metadata of the workbook instance
        self._update_meta(trace, app_dict)

        # Check if sheet is already active
        if self.active_sheet_title == worksheet:
            worksheet_obj = self.active_sheet

        # Set the requested sheet active
        else:
            # Check if sheet exists
            if worksheet in self.wb_sheets:
                self.activate_worksheet(trace, app_dict, worksheet)
                worksheet_obj = self.active_sheet

            elif not worksheet:
                worksheet_obj = self.active_sheet
                worksheet = self.active_sheet_title

            # The requested sheet was not found
            else:
                # Could not find the requested worksheet.
                raise ValueError(
                    f'{app_dict["loc"]["morpy"]["read_cells_not_found"]}\n'
                    f'{app_dict["loc"]["morpy"]["read_cells_file"]}: {self.wb_path}\n'
                    f'{app_dict["loc"]["morpy"]["read_cells_sht"]}: {worksheet}\n'
                    f'{app_dict["loc"]["morpy"]["read_cells_av_sheets"]}: {self.wb_sheets}'
                )

        if not cell_range:
            cell_range = [
                f'{openpyxl.utils.cell.get_column_letter(self.active_sheet.min_column)}{self.active_sheet.min_row}:'
                f'{openpyxl.utils.cell.get_column_letter(self.active_sheet.max_column)}{self.active_sheet.max_row}'
            ]

        # Autoformat cell reference(s)
        cl_dict = self._cell_ref_autoformat(trace, app_dict, cell_range)["cl_dict"]

        # Loop through all the cells and read them
        # TODO implement gui here, total progress is len(cl_dict.keys())
        for cl in cl_dict:
            cell_obj = worksheet_obj[cl]

            if cell_styles:
                cl_dict[cl] =  {
                    "value": cell_obj.value,
                    "comment": {
                        "text": cell_obj.comment.text if cell_obj.comment else None,
                        "author": cell_obj.comment.author if cell_obj.comment else None,
                    },
                    "format": cell_obj.number_format,
                    "font": {
                        "name": cell_obj.font.name,
                        "bold": cell_obj.font.bold,
                        "italic": cell_obj.font.italic,
                        "vertical align": cell_obj.font.vertAlign,
                        "underline": cell_obj.font.underline,
                        "strike": cell_obj.font.strike,
                        "size": cell_obj.font.sz,
                        "color": cell_obj.font.color.rgb if cell_obj.font.color else None,
                    },
                    "background": {
                        "fill type": cell_obj.fill.fill_type,
                        "start color": cell_obj.fill.start_color.rgb if cell_obj.fill.start_color else None,
                        "end color": cell_obj.fill.end_color.rgb if cell_obj.fill.end_color else None,
                    },
                    "border": {
                        "top": {
                            "style": cell_obj.border.top.style if cell_obj.border.top else None,
                            "color": cell_obj.border.top.color.rgb if cell_obj.border.top.color else None,
                        },
                        "bottom": {
                            "style": cell_obj.border.bottom.style if cell_obj.border.bottom else None,
                            "color": cell_obj.border.bottom.color.rgb if cell_obj.border.bottom.color else None,
                        },
                        "left": {
                            "style": cell_obj.border.left.style if cell_obj.border.left else None,
                            "color": cell_obj.border.left.color.rgb if cell_obj.border.left.color else None,
                        },
                        "right": {
                            "style": cell_obj.border.right.style if cell_obj.border.right else None,
                            "color": cell_obj.border.right.color.rgb if cell_obj.border.right.color else None,
                        },
                    },
                    "alignment": {
                        "horizontal": cell_obj.alignment.horizontal,
                        "vertical": cell_obj.alignment.vertical,
                        "text rotation": cell_obj.alignment.textRotation,
                        "wrap text": cell_obj.alignment.wrapText,
                        "shrink to fit": cell_obj.alignment.shrinkToFit,
                        "indent": cell_obj.alignment.indent,
                    },
                    "protection": {
                        "locked": cell_obj.protection.locked,
                        "hidden": cell_obj.protection.hidden,
                    },
                }
            else:
                cl_dict[cl] =  {
                    "value": cell_obj.value,
                }

        # The worksheet was read from.
        log(trace, app_dict, "debug",
            lambda: f'{app_dict["loc"]["morpy"]["read_cells_read"]}\n'
                f'{app_dict["loc"]["morpy"]["read_cells_file"]}: {self.wb_path}\n'
                f'{app_dict["loc"]["morpy"]["read_cells_sht"]}: {worksheet}\n'
                f'{app_dict["loc"]["morpy"]["read_cells_cls"]}: {cell_range}')

        return{
            'cl_dict' : cl_dict
            }


    @morpy_wrap
    def write_ranges(self, trace: dict, app_dict: dict, worksheet: str=None, cell_range: list=None,
                     cell_writes: list=None, fill_range: bool=False, style_default: bool=False,
                     save_workbook: bool=True, close_workbook: bool=False) -> None:
        r"""
        Writes data into cells of an Excel workbook. OpenPyXL documentation:
        https://openpyxl.readthedocs.io/en/stable/api/openpyxl.cell.cell.html#openpyxl.cell.cell.Cell
        https://openpyxl.readthedocs.io/en/3.1.3/styles.html

        :param trace: Operation credentials and tracing information.
        :param app_dict: The morPy global dictionary containing app configurations.
        :param worksheet: Name of the worksheet, where the cell is located. If None, the
            active sheet is addressed.
        :param cell_range: The cell or range of cells to write. Accepted formats:
            - not case-sensitive
            - Single cell: ["A1"]
            - Range of cells: ["A1:ZZ1000"]
            - Ranges of cells: ["A1:ZZ1000", "c2:fl342"]
        :param cell_writes: List of cell content dictionaries to be written consecutively. If the list is shorter
            than the amount of cells in the range, it will stop writing to cells and finish the operation. The
            dictionaries do not need to contain all possible keys, only assign the cell attributes/values needed.
            See example for a display of usage. Following is an example of a single complete cell write:
                cl_list = [
                    {"value" : "Data in Cell",
                    "comment" : {
                        "text" : "This is a comment",
                        "author" : "Mr. Author Man",
                    },
                    "format" : "General",           # Options: General|Number|Currency|Accounting|Date|Time|Percentage
                                                    # > Fraction|Scientific|Text|custom strings
                    "font" : {
                        "name" : "Calibri",
                        "bold" : True,
                        "italic" : True,
                        "vertical align" : None,    # Options: None|superscript|subscript
                        "underline" : None,         # Options: None|single|double|singleAccounting|doubleAccounting
                        "strike" : False,
                        "size" : 14,
                        "color" : "D1760C",
                    },
                    "background" : {
                        "fill type" : None,         # Options: None|solid|darkGrid|darkTrellis|lightDown|lightGray
                                                    # > lightGrid|lightHorizontal|lightTrellis|lightUp
                                                    # > lightVertical|mediumGray
                        "start color" : "307591",
                        "end color" : "67CCCC",
                    },
                    "border" : {
                        "edge" : "outline",         # Options: left|right|top|bottom|diagonal|outline
                                                    # > vertical|horizontal
                        "style" : None,             # Options: None|dashDot|dashDotDot|dashed|dotted|double|hair|medium
                                                    # > medium|mediumDashDot|mediumDashDotDot|mediumDashed|slantDashDot
                                                    # > thick|thin
                        "color" : "2A7F7F",
                        "diagonal direction" : 0,   # Options: 0 (no diagonal) | 1 (downwards) | 2 (upwards)
                    },
                    "alignment" : {
                        "horizontal" : "general",   # Options: general|left|center|right|fill|justify|centerContinuous
                                                    # > distributed
                        "vertical" : "center",      # Options: top|center|bottom|justify|distributed
                        "text rotation" : 0,
                        "wrap text" : False,
                        "shrink to fit" : False,
                        "indent" : 0,
                    },
                    "protection" : {
                        "locked" : False,
                        "hidden" : False,
                    },
                },
                { ... }]
        :param fill_range: If True and if the cell_writes list is shorter than the amount
            of cells in the range, it will continue writing the values from beginning until the end
            of the range.
        :param style_default: If True, styles/attributes of cells will be reset to default. If False,
            the styles/attributes of the original cell will be preserved.
        :param save_workbook: If True, saves the workbook after the changes.
        :param close_workbook: If True, closes the workbook.

        :example:
            import morPy
            wb_path = "C:\my_file.xlsx"
            wb = morPy.XlWorkbook(trace, app_dict, wb_path)

            w_sh = "Sheet1"
            range = ["A1", "B2:C3", "B1:ZZ5"]
            cell_writes = []
            cell_writes.append(
                {"value": "Example"}
            )
            cell_writes.append({
                "value" : r"=1+2",
                "font" : {
                    "color" : "D1760C",
                    },
            })
            wb.write_cells(
                trace, app_dict,
                worksheet=w_sh,
                cell_range=cl_rng,
                cell_writes=cell_writes,
                save_workbook=True, close_workbook=True
            )
        """

        if not worksheet:
            self._update_meta(trace, app_dict, minimal=True)
            worksheet_obj = self.active_sheet
        else:
            self.activate_worksheet(trace, app_dict, worksheet)
            worksheet_obj = self.active_sheet

        if cell_range:
            cl_dict = self._cell_ref_autoformat(trace, app_dict, cell_range=cell_range)["cl_dict"]
            cell_keys = list(cl_dict.keys())
            write_count = len(cell_writes)

            for i, cell_key in enumerate(cell_keys):
                cell_obj = worksheet_obj[cell_key]
                write_data = cell_writes[i % write_count] if fill_range else cell_writes[i] if i < write_count else None

                # Finish cell writes conditionally
                if not write_data:
                    break

                # Reset styles if requested
                if style_default:
                    cell_obj.style = 'Normal'

                # Get cell styles to be merged with cell_writes
                cell_data = self.read_cells(trace, app_dict, cell_range=[cell_key],
                                                worksheet=self.active_sheet_title)

                # Merge current styles with the ones provided in write_data
                current_styles = cell_data["cl_dict"]  # e.g. {"A2": {"value":..., "font":...}}
                cell_styles = current_styles.get(cell_key, {})  # e.g. {"value":..., "font":...}

                # Merge in any new style info from write_data
                for style_key, style_val in write_data.items():
                    if style_key in cell_styles and isinstance(style_val, dict):
                        # Merge with existing dictionary (e.g. "font", "alignment", "border")
                        cell_styles[style_key].update(style_val)
                    else:
                        # Overwrite or add a new key (e.g. "value")
                        cell_styles[style_key] = style_val

                # Apply to the real cell:
                if "value" in cell_styles:
                    cell_obj.value = cell_styles["value"]
                if "comment" in cell_styles and cell_styles["comment"]:
                    from openpyxl.comments import Comment
                    cell_obj.comment = Comment(
                        text=cell_styles["comment"].get("text", ""),
                        author=cell_styles["comment"].get("author", ""),
                    )
                if "font" in cell_styles:
                    cell_obj.font = Font(**cell_styles["font"])
                if "background" in cell_styles:
                    cell_obj.fill = PatternFill(
                        fill_type=cell_styles["background"]["fill type"],
                        start_color=cell_styles["background"]["start color"],
                        end_color=cell_styles["background"]["end color"],
                    )
                if "alignment" in cell_styles:
                    cell_obj.alignment = Alignment(**cell_styles["alignment"])
                if "protection" in cell_styles:
                    cell_obj.protection = Protection(**cell_styles["protection"])

        elif not cell_range:
            # Missing cell range. Skipped writing to cells.
            raise ValueError(f'{app_dict["loc"]["morpy"]["write_cells_no_range"]}')

        # Cells written to.
        log(trace, app_dict, "debug",
            lambda: f'{app_dict["loc"]["morpy"]["write_cells_done"]}\n'
                    f'{app_dict["loc"]["morpy"]["write_cells_range"]}: {cell_range}')

        if save_workbook:
            self.save_workbook(trace, app_dict, close_workbook=close_workbook)


    @morpy_wrap
    def get_table_attributes(self, trace: dict, app_dict: dict, table: str) -> dict:
        r"""
        Retrieves all attributes of an MS Excel table. OpenPyXL documentation:
        https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.table.html#openpyxl.worksheet.table.Table

        :param trace: Operation credentials and tracing information.
        :param app_dict: The morPy global dictionary containing app configurations.
        :param table: Name of the table to be analyzed.

        :return: dict
            table_attr: List of the table's attributes.

        :example:
            import morPy
            wb_path = "C:\my_file.xlsx"
            wb = morPy.XlWorkbook(trace, app_dict, wb_path)

            table = "tbl_finances"
            table_attr = wb.get_table_attributes(trace, app_dict, table)["table_attr"]
        """

        # Update metadata of the workbook instance
        self._update_meta(trace, app_dict, minimal=True)

        # Inquire the according worksheet of the table
        worksheet = self.tables_sheets[table]

        # Get all values of the table
        table_data = self.wb_obj[worksheet].tables.values()
        table_data = openpyxl_table_data_dict(trace, app_dict, table_data, table)
        table_attr = table_data["table_attr"]

        # Retrieved all values of an MS Excel table.
        log(trace, app_dict, "debug",
            lambda: f'{app_dict["loc"]["morpy"]["get_table_attributes_retrieved"]}\n'
                    f'{app_dict["loc"]["morpy"]["get_table_attributes_path"]}: {self.wb_path}\n'
                    f'{app_dict["loc"]["morpy"]["get_table_attributes_sheet"]}: {worksheet}\n'
                    f'{app_dict["loc"]["morpy"]["get_table_attributes_table"]}: {table}')

        return{
            'table_attr' : table_attr,
            }


@morpy_wrap
def openpyxl_table_data_dict(trace: dict, app_dict: dict, table_data: object, table: str) -> dict:
    r"""
    Converts the interface of an OpenPyXL data book into a dictionary containing
    all attributes of the specified table. This function is a helper and is
    typically called by `get_all_tables_attributes` to improve on the OpenPyXL
    API.

    :param trace: Operation credentials and tracing information.
    :param app_dict: The morPy global dictionary containing app configurations.
    :param table_data: The data book object as generated by OpenPyXL.
    :param table: Name of the table to be analyzed.

    :return: dict
        table_attr: List containing all attributes of the OpenPyXL data book.
    """

    table_data = f'{table_data}'
    table_item = ""
    table_attr = []

    # Search for regular expressions in the data book to extract only the
    # relevant Part.

    # 1. Purge all whitespace characters to make regex easier and more precise
    table_data = common.regex_replace(trace, app_dict, table_data, r'\s', '')

    # 2. Split the data-book into a list of distinct table attributes
    delimiter = '<openpyxl.worksheet.table.Table object>'
    table_data_list = common.regex_split(trace, app_dict, table_data, delimiter)

    # 3. Iterate through the list in search for the table and delete
    # elements not associated with it
    pattern = f'\'{table}\''

    for table_item in table_data_list:

        result = common.regex_find1st(trace, app_dict, table_item, pattern)

        if result:
            break

    # 4. Replace the comma at the end of the string, if there is any
    table_item = common.regex_replace(trace, app_dict, table_item, ',$', '')

    # 5. Add the first delimiter and reinsert some spaces for compatibility
    # with OpenPyXL
    table_item = f'<openpyxl.worksheet.table.Table object>{table_item}'
    table_item = common.regex_replace(trace, app_dict, table_item, 'object>', ' object>')

    # 6. Split the string into different sections
    table_attr = common.regex_split(trace, app_dict, table_item, ',')

    # Converted an OpenPyXL data-book into a list specific to the attributes of the MS Excel table.
    log(trace, app_dict, "debug",
        lambda: f'{app_dict["loc"]["morpy"]["openpyxl_table_data_dict_conv"]}\n'
                f'{app_dict["loc"]["morpy"]["openpyxl_table_data_dict_tbl"]}: {table}\n'
                f'{app_dict["loc"]["morpy"]["openpyxl_table_data_dict_attr"]}:\n{table_attr}')

    return{
        'table_attr' : table_attr
        }