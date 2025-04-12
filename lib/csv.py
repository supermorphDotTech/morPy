r"""
morPy Framework by supermorph.tech
https://github.com/supermorphDotTech

Author:     Bastian Neuwirth
Descr.:     Module of operations concerning csv-files.
"""

import lib.fct as morpy_fct
import lib.xl as xl
from lib.decorators import metrics, log
from lib.common import ProgressTracker

import sys
from openpyxl.utils.cell import get_column_letter

@metrics
def csv_read(morpy_trace: dict, app_dict: dict, src_file_path: str=None, delimiter: str=None,
             print_csv_dict: bool=False, log_progress: bool=False, progress_ticks: float=None, gui=None) -> dict:
    r"""
    This function reads a csv file and returns a dictionary of
    dictionaries. The header row, first row of data and delimiter
    is determined automatically.

    :param morpy_trace: Operation credentials and tracing
    :param app_dict: morPy global dictionary containing app configurations
    :param src_file_path: Path to the csv file.
    :param delimiter: Delimiters used in the csv. None = Auto detection
    :param print_csv_dict: If true, the csv_dict will be printed to console. Should only
        be used for debugging with small example csv files. May take very long to complete.
    :param log_progress: If True, logs the progress.
    :param progress_ticks: Percentage of total to log the progress. I.e. at ticks=10.7 at every
        10.7% progress exceeded the exact progress will be logged. If None or greater 100, will default to 10.
        If gui is not None, may overwrite this setting.
    :param gui: User Interface reference. Automatically referenced by morPy.ProgressTrackerTk()

    :return: dict
        morpy_trace: Operation credentials and tracing
        check: Indicates whether the function ended without errors
        csv_dict: Dictionary containing all tags. The line numbers of data are
            the keys of the parent dictionary, and the csv header consists of
            the keys of every sub-dictionary.
            Pattern:
                {DATA1 :
                    delimiter : [str]
                    header : [tuple] (header(1), header(2), ...)
                    columns : [int] header columns
                    rows : [int] number of rows in data
                    ROW1 : [dict]
                        {header(1) : data(1, 1),
                        header(2) : data(2, 1),
                        ...}
                    ROW2 : [dict]
                        {header(1) : data(1, 2),
                        header(2) : data(2, 2),
                        ...}
                DATA2 : ...}

    :example:
        src_file_path = 'C:\my_file.csv'
        delimiter = '\",\"'
        csv = csv_read(morpy_trace, app_dict, src_file_path, delimiter)
        csv_dict = csv["csv_dict"]
        csv_header1 = csv["csv_dict"]["DATA1"]["header"]
        print(f'{csv_header1}')
    """

    # Define operation credentials (see init.init_cred() for all dict keys)
    module: str = 'lib.csv'
    operation: str = 'csv_read(~)'
    morpy_trace: dict = morpy_fct.tracing(module, operation, morpy_trace)

    csv_copy_dict: dict         = {}
    csv_dict: dict              = {}
    header: tuple               = ()
    data_table: str             = ''
    columns: int                = 0
    header_row: int             = -1
    data_rows: int              = 0         # Sum of data rows in a data table
    r_data: int                 = 0         # Row counter for file processing
    delimiter_identified: bool  = False     # True, if delimiter was successfully identified
    header_cnt_delimiters: int  = -1        # Number of determined columns in header
    data_subdict_nr: int        = 0
    data_cnt_row: int           = 0         # Counter for printing the data row sub-dictionary
    csv_read_progress           = None
    gui_msg_row: str            = ''

    try:
        # Started processing CSV-file.
        log(morpy_trace, app_dict, "info",
        lambda: f'{app_dict["loc"]["morpy"]["csv_read_start"]}\n'
                f'{app_dict["loc"]["morpy"]["csv_read_file_path"]}: {src_file_path}')

        # Delimiter auto-detection
        delimiters = (delimiter,) if delimiter else ('";"', '","', ';', ',', '"\t"', '\t', '":"', ':')

        src_file_dict = morpy_fct.pathtool(src_file_path)
        src_file_path = src_file_dict["out_path"]
        src_file_isfile = src_file_dict["is_file"]
        src_file_exists = src_file_dict["file_exists"]
        src_file_ext = src_file_dict["file_ext"]

        if src_file_isfile and src_file_exists and src_file_ext == ".csv":
            r_tot = 0 # Total row counter
            # Import the csv file as is and determine length
            with open(src_file_path, 'r') as csv_file:
                for line in csv_file:
                    r_tot +=1
                    csv_copy_dict[f'{r_tot}'] = line

            # Track progress
            if log_progress:
                prog_total = len(delimiters) * r_tot
                if gui:
                    gui.begin_stage(morpy_trace, app_dict,
                                    stage_limit=prog_total,
                                    headline_stage=app_dict["loc"]["morpy"]["csv_read_start"],
                                    detail_description=app_dict["loc"]["morpy"]["csv_read_stage"],
                                    ticks=25)
                    gui_msg_row = app_dict["loc"]["morpy"]["csv_read_row"]
                else:
                    csv_read_progress = ProgressTracker(morpy_trace, app_dict,
                        description=app_dict["loc"]["morpy"]["csv_read_stage"], total=prog_total, ticks=progress_ticks)

            # Determine delimiters, header and data rows
            for d in delimiters:
                r_det = 0 # Row counter for delimiter check
                for row, line in csv_copy_dict.items():
                    r_det +=1

                    # Progress Tracking
                    if log_progress:
                        if gui:
                            gui.update_text(morpy_trace, app_dict, detail_description=f'{gui_msg_row} {r_det}')
                            gui.update_progress(morpy_trace, app_dict, current=r_det)
                        else:
                            csv_read_progress.update(morpy_trace, app_dict, current=r_det)

                    line = line.rstrip()
                    if d in line:
                        # Check, if header needs to be found
                        if header_row < 0:
                            header_row = r_det
                            header = tuple(map(lambda x: x.strip('"\''), line.split(d)))
                            header_cnt_delimiters = len(header)
                            columns = header_cnt_delimiters + 1
                        # If header was assigned, check for data row
                        else:
                            data = tuple(map(lambda x: x.strip('"\''), line.split(d)))
                            data_cnt_delimiters = len(data)
                            data_row = r_det
                            r_data += 1
                            # If header and data got the same amount of columns, data table found.
                            if header_cnt_delimiters == data_cnt_delimiters:
                                # Check, if new data table
                                if data_row == header_row + 1:
                                    delimiter_identified: bool = True
                                    data_subdict_nr += 1
                                    data_table = f'DATA{data_subdict_nr}'
                                    csv_dict[f'{data_table}'] = {}
                                    csv_dict[f'{data_table}']["delimiter"] = d
                                    csv_dict[f'{data_table}']["header"] = header
                                    csv_dict[f'{data_table}']["columns"] = columns
                                    csv_dict[f'{data_table}']["rows"] = data_rows
                                csv_dict[f'{data_table}'][f'{r_data}'] = dict(zip(header, data))
                                data_rows += 1
                                csv_dict[f'{data_table}']["rows"] = data_rows
                            # Set current line as header for next iteration of determination
                            else:
                                header_row = r_det
                                header_cnt_delimiters = data_cnt_delimiters
                                header = tuple(map(lambda x: x.strip(), line.split(d)))
                                columns = header_cnt_delimiters + 1
                                data_rows = 0
                                r_data = 0
                    else:
                        header_row = -1
                        header_cnt_delimiters = -1
                        data_rows = 0
                        r_data = 0
                        header = ()

            # Process CSV into dictionary
            if delimiter_identified:
                # CSV file processed. Dictionary contains ## rows.
                log(morpy_trace, app_dict, "info",
                lambda: f'{app_dict["loc"]["morpy"]["csv_read_done"]}: {data_rows}')

                # Print csv_dict to console for debugging.
                if print_csv_dict:
                    # Start printing returned dictionary to console.
                    log(morpy_trace, app_dict, "info",
                    lambda: f'{app_dict["loc"]["morpy"]["csv_read_done"]}: {data_rows}')

                    for data_table_name in csv_dict.keys():
                        print(f'{0*" "}{data_table_name}: {{')
                        for meta, meta_val in csv_dict[data_table_name].items():
                            # Check, if meta holds data and therefore a dictionary
                            if isinstance(meta_val, dict):
                                data_cnt_row +=1
                                print(f'{3*" "}ROW{data_cnt_row}: {{')
                                for dat_col, dat_row in meta_val.items():
                                    print(f'{6*" "}{dat_col} : {dat_row}')
                                print(f'{3*" "}}}')
                            else:
                                print(f'{3*" "}{meta} : {meta_val}')
                        print(f'{0*" "}}}\n')
                        data_cnt_row = 0

                    # Finished printing returned dictionary to console.
                    log(morpy_trace, app_dict, "info",
                    lambda: f'{app_dict["loc"]["morpy"]["csv_read_done"]}: {data_rows}')

            else:
                # Delimiters could not be determined or data is corrupted. No return dictionary created.
                log(morpy_trace, app_dict, "warning",
                lambda: f'{app_dict["loc"]["morpy"]["csv_read_no_return"]}\n'
                        f'{app_dict["loc"]["morpy"]["csv_read_file_path"]}: {src_file_path}')

        else:
            # File does not exist or is not a CSV file.
            log(morpy_trace, app_dict, "warning",
            lambda: f'{app_dict["loc"]["morpy"]["csv_read_not_done"]}\n'
                    f'{app_dict["loc"]["morpy"]["csv_read_file_exist"]}: {src_file_exists}\n'
                    f'{app_dict["loc"]["morpy"]["csv_read_isfile"]}: {src_file_isfile}\n'
                    f'{app_dict["loc"]["morpy"]["csv_read_file_ext"]}: {src_file_ext}\n'
                    f'{app_dict["loc"]["morpy"]["csv_read_file_path"]}: {src_file_path}')

        check: bool = True

    except Exception as e:
        from lib.exceptions import MorPyException
        raise MorPyException(morpy_trace, app_dict, e, sys.exc_info()[-1].tb_lineno, "error")

    return{
        'morpy_trace' : morpy_trace,
        'check' : check,
        'csv_dict' : csv_dict
        }

@metrics
def csv_dict_to_excel(morpy_trace: dict, app_dict: dict, xl_path: str=None, overwrite: bool=False,
                      worksheet: str=None, close_workbook: bool=False, csv_dict: dict=None,
                      log_progress: bool=False, progress_ticks: float=None) -> dict:
    r"""
    This function takes da dictionary as provided by csv_read() and saves it as an MS Excel file.
    The csv_dict however may be evaluated and processed prior to executing csv_dict_to_excel().
    The file will be saved automatically, but closing the workbook is optional.

    :param morpy_trace: Operation credentials and tracing
    :param app_dict: morPy global dictionary containing app configurations
    :param xl_path: Path to the target MS Excel file.
    :param overwrite: If True, an existing MS Excel file may be overwritten.
    :param worksheet: Name of the worksheet, where the cell is located. If None, the
        active sheet is addressed.
    :param close_workbook: If True, closes the workbook.
    :param csv_dict: Dictionary containing all tags. The line numbers of data are
        the keys of the parent dictionary, and the csv header consists of
        the keys of every sub-dictionary.
        Pattern:
            {DATA1 :
                delimiter : [str]
                header : [tuple] (header(1), header(2), ...)
                columns : [int] header columns
                rows : [int] number of rows in data
                ROW1 : [dict]
                    {header(1) : data(1, 1),
                    header(2) : data(2, 1),
                    ...}
                ROW2 : [dict]
                    {header(1) : data(1, 2),
                    header(2) : data(2, 2),
                    ...}
            DATA2 : ...}
    :param log_progress: If True, logs the progress.
    :param progress_ticks: Percentage of total to log the progress. I.e. at ticks=10.7 at every
        10.7% progress exceeded the exact progress will be logged. If None or greater 100, will default to 10.

    :return: dict
        morpy_trace: Operation credentials and tracing
        check: Indicates whether the function ended without errors
        wb_obj: Returns None, if the object was closed. Else returns an instance of "xl.XlWorkbook()".
            Used to delete the reference to an instance.

    :example:
        src_file_path = 'C:\my.csv'
        delimiter = '","'
        csv = morPy.csv_read(morpy_trace, app_dict, src_file_path, delimiter)

        # ... process data in csv["csv_dict"] ...

        target_path = 'C:\my.xlsx'
        wb_sht = 'Sheet1'
        wb = csv_dict_to_excel(morpy_trace, app_dict, csv_dict=csv["csv_dict"], xl_path=target_path,
            overwrite=True, worksheet=wb_sht)["wb_obj"]

        # ... modify 'C:\my.xlsx' ...

        # Save and close workbook
        wb.close_workbook(mpy_trac, app_dict, save_workbook=True, close_workbook=True)
    """

    # Define operation credentials (see init.init_cred() for all dict keys)
    module: str = 'lib.csv'
    operation: str = 'csv_dict_to_excel(~)'
    morpy_trace: dict = morpy_fct.tracing(module, operation, morpy_trace)

    check: bool = False
    xl_exists: bool = False
    xl_write: bool = False
    wb_obj = None
    progress = None

    try:
        # Writing data to MS Excel file.
        log(morpy_trace, app_dict, "info",
            lambda: f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_start"]}\n'
                    f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_path"]}: {xl_path}\n'
                    f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_overwrite"]}: {overwrite}')

        # Log the progress
        if log_progress:
            total_prog_count = 0
            for d_block in csv_dict.keys():
                row_cnt = csv_dict[d_block].get("rows")
                if row_cnt:
                    total_prog_count += csv_dict[d_block].get("rows")
                else:
                    log_progress = False

                    # Missing row count in csv_dict. Skipping progress logging.
                    log(morpy_trace, app_dict, "warning",
                        lambda: f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_prog_fail"]}')

            # Instantiate progress logging
            progress = ProgressTracker(
                morpy_trace, app_dict, description=f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_prog_descr"]}',
                total=total_prog_count, ticks=progress_ticks
            )

        # Check target file path
        if xl_path:
            xl_path_eval = morpy_fct.pathtool(in_path=xl_path)
            xl_exists = xl_path_eval["file_exists"]
            xl_valid = True
        else:
            xl_valid = False

        # Evaluate, if write to file is allowed and path is valid
        if xl_valid and (not xl_exists or xl_exists and overwrite):
            xl_write = True

        # If we have a valid path and a dictionary with data, attempt to write
        if xl_write and csv_dict:
            # Instantiate workbook API
            wb = xl.XlWorkbook(morpy_trace, app_dict, xl_path, create=not xl_exists)

            row_index = 1  # Starting row

            # Loop over each data block in csv_dict
            for data_block_key, data_block_val in csv_dict.items():

                # -- Validate the block --
                if not isinstance(data_block_val, dict):
                    continue
                if "header" not in data_block_val or not data_block_val["header"]:
                    continue

                # 1) Write the header row
                header_tuple = data_block_val["header"]
                header_length = len(header_tuple)

                # Build a range for the header from e.g. A{row_index}:E{row_index}
                # If header_length is 5, columns = A, B, C, D, E
                # Example range: ["A{row_index}:E{row_index}"]
                first_col_letter = get_column_letter(1)
                last_col_letter = get_column_letter(header_length)
                header_range = [f"{first_col_letter}{row_index}:{last_col_letter}{row_index}"]

                # Construct a list of dictionaries representing each header cell.
                # Each cell’s "value" is the corresponding column name in the header.
                # Styles are applied conditionally.
                cell_writes_for_header = []
                for col_header in header_tuple:
                    cell_writes_for_header.append({
                        "value": col_header
                    })

                # Write the header row
                wb.write_ranges(
                    morpy_trace,
                    app_dict,
                    worksheet=worksheet,
                    cell_range=header_range,
                    cell_writes=cell_writes_for_header,
                    fill_range=False,                       # no re-cycling of style/data in the same row
                    style_default=False,                    # keep default or user-specified style
                    save_workbook=False,                    # we'll do a final save after everything
                    close_workbook=False,
                )

                row_index += 1  # move to the next row to start writing data

                # 2) Write all data rows
                # Convert string keys to int, sort them if needed
                # The block might have items like "1", "2", ... representing row1, row2...
                # We'll gather them in sorted order
                row_keys = []
                for k in data_block_val.keys():
                    # skip the block-level keys
                    if k in ("delimiter", "header", "columns", "rows"):
                        continue
                    try:
                        row_keys.append(int(k))
                        if log_progress:
                            progress.update(morpy_trace, app_dict)
                    except (ValueError, TypeError):
                        pass # Ignore keys that cannot be converted to integers
                row_keys.sort()

                for row_key in row_keys:
                    # data_dict is e.g. {'Name': 'Alice', 'Age': '42'}
                    data_dict = data_block_val[str(row_key)]

                    # We build a range for the row e.g. A{row_index}:E{row_index}
                    row_range = [f"{first_col_letter}{row_index}:{last_col_letter}{row_index}"]

                    # Now build a list of each cell’s data in the order of the header
                    cell_writes_for_data = []
                    for col_header in header_tuple:
                        # fetch the cell data (string, number, whatever)
                        cell_value = data_dict.get(col_header, "")
                        cell_writes_for_data.append({"value": cell_value})

                    wb.write_ranges(
                        morpy_trace,
                        app_dict,
                        worksheet=worksheet,
                        cell_range=row_range,
                        cell_writes=cell_writes_for_data,
                        fill_range=False,
                        style_default=False,
                        save_workbook=False,
                        close_workbook=False,
                    )

                    row_index += 1  # next row

                # 3) Skip one row after finishing this block
                row_index += 1

            # Save and optionally close the workbook
            wb_obj = wb.save_workbook(
                morpy_trace,
                app_dict,
                close_workbook=close_workbook,
            )["wb_obj"]

            check: bool = True

        if xl_exists and overwrite:
            # MS Excel file exists. Overwritten.
            log(morpy_trace, app_dict, "debug",
                lambda: f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_xl_overwrite"]}\n'
                        f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_path"]}: {xl_path}\n'
                        f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_overwrite"]}: {overwrite}')

        elif xl_exists and not overwrite:
            # MS Excel file exists. Operation skipped.
            log(morpy_trace, app_dict, "debug",
                lambda: f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_xl_not_overwrite"]}\n'
                        f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_path"]}: {xl_path}'
                        f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_overwrite"]}: {overwrite}')

        if not xl_path:
            # Missing path to MS Excel file. Operation skipped.
            log(morpy_trace, app_dict, "warning",
            lambda: f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_missing_xl"]}\n'
                    f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_path"]}: {xl_path}')

        if not xl_valid:
            # Invalid path to MS Excel file. Operation skipped.
            log(morpy_trace, app_dict, "warning",
            lambda: f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_invalid_xl"]}\n'
                    f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_path"]}: {xl_path}')

        elif not csv_dict:
            # Missing data book from csv file. operation skipped.
            log(morpy_trace, app_dict, "warning",
            lambda: f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_missing_data"]}\n'
                    f'{app_dict["loc"]["morpy"]["csv_dict_to_excel_data"]}: {csv_dict}')

    except Exception as e:
        from lib.exceptions import MorPyException
        raise MorPyException(morpy_trace, app_dict, e, sys.exc_info()[-1].tb_lineno, "error")

    return{
        'morpy_trace' : morpy_trace,
        'check' : check,
        'wb_obj' : wb_obj,
        }